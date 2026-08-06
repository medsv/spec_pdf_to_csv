import os
import tempfile
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Импортируем вашу библиотеку
from libs.spec_pdf_to_csv import spec_pdf_to_row_list

# Настройки страницы
st.set_page_config(
    page_title="Конвертер спецификаций PDF в XLSX",
    page_icon="📄",
    layout="centered",
)

# Путь к шаблону спецификации
TEMPLATE_PATH = Path("templates") / "Шаблон_спецификации_РД.xlsx"

st.title("📄 Конвертер спецификаций рабочей документации (PDF -> XLSX)")
st.markdown("Загрузите PDF-файл спецификации для её перевода в формат XLSX.")


def build_xlsx_bytes(spec, pdf_stem, template_path):
    """Формирует xlsx-файл спецификации на базе шаблона и возвращает его в виде байтов.

    spec: список строк спецификации (список списков).
    pdf_stem: имя исходного pdf-файла без расширения, используется в качестве заголовка A1.
    template_path: путь к xlsx-файлу шаблона спецификации.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон спецификации не найден: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    if ws is None:
        raise ValueError("Не удалось получить активный лист из шаблона")

    ws["A1"] = pdf_stem
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row_values in enumerate(spec, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# 1. Загрузка файла
pdf_file = st.file_uploader(
    "Выберите PDF файл, нажав Upload, или перетащите его сюда мышкой ", type=["pdf"]
)

if pdf_file is not None:
    st.info(f"Загружен файл: **{pdf_file.name}** ({pdf_file.size / 1024:.1f} КБ)")

    # 2. Кнопка запуска обработки
    if st.button("🚀 Извлечь данные и сформировать XLSX", type="primary"):
        with st.spinner("Идёт анализ PDF ... Пожалуйста, подождите."):
            pdf_path = None
            try:
                # Сохраняем PDF во временный файл (нужен pymupdf)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                    tmp_pdf.write(pdf_file.getvalue())
                    pdf_path = tmp_pdf.name

                spec = spec_pdf_to_row_list(pdf_path)
                xlsx_bytes = build_xlsx_bytes(spec, Path(pdf_file.name).stem, TEMPLATE_PATH)
                xlsx_name = Path(pdf_file.name).with_suffix(".xlsx").name

                st.success("✅ Файл успешно обработан!")

                # 3. Кнопка скачивания
                st.download_button(
                    label="📥 Скачать Excel файл",
                    data=xlsx_bytes,
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

            except Exception as e:
                st.error(f"❌ Произошла ошибка при обработке файла: {e}")
                st.exception(e)

            finally:
                # 4. Очистка временного PDF-файла
                if pdf_path:
                    try:
                        os.remove(pdf_path)
                    except OSError:
                        pass

else:
    st.warning("Пожалуйста, загрузите PDF файл для начала работы.")

st.markdown(
    """
    <hr>
    <p style="text-align: left; color: gray;">
    <small>
    2026, С.В. Медведев, engpython@yandex.ru
    </small>
    </p>
    """,
    unsafe_allow_html=True,
)
