import pymupdf
from copy import copy
from io import BytesIO
from openpyxl import load_workbook
from libs.utils import normalize_row, gost_spec_title, correct_row
from openpyxl.styles import PatternFill


def pdf_spec_to_row_list(pdf_path):
    """Извлекает из pdf-файла спецификации список строк спецификации."""
    with pymupdf.open(pdf_path) as doc:
        return parse_spec(doc)


def parse_spec(doc):
    spec = []
    spec_row_count = 0
    restored_rows: dict[int, list[int]] = {}
    prev_spec_line = None
    first_table = True
    template_cols_count = 9  # по умолчанию табдица по ГОСТ
    for page in doc:
        tabs = page.find_tables()  # locate and extract any tables on page
        if not tabs:
            continue
        for tab in tabs:
            in_spec = False  # не дошёл до спецификации
            lines = tab.extract()
            for line in lines:
                if len(line) < template_cols_count: continue
                #print(line)
                if not in_spec:
                    #if "Примечание" in line or "Код продукции" in line:  # шапка таблицы спецификации
                    if any("приме" in str(s).lower().strip() for s in line) or \
                            all(str(dig) in line for dig in range(1,10)): # шапка таблицы спецификации
                        template_cols_count = 9
                        #if any("kks" in str(s).lower().strip() for s in line): template_cols_count = 10
                        spec_line_cols_count = len(line)  # количество столбцов в pdf-таблице представления спецификации
                        pattern = detect_pattern(line)  # шаблон таблицы спецификации
                        spec_col_count = len(pattern)
                        if spec_col_count != template_cols_count:
                            continue
                            #raise ValueError(f"В спецификации должно быть 9 столбцов, а не {spec_col_count}")
                        if first_table:  # только для первой шапки таблицы
                            spec.append(gost_spec_title())  # добавляем в спецификацию шапку по ГОСТ 21.110-2013
                            spec_row_count += 1
                        in_spec = True # внутри спецификации
                        first_table = False
                        continue
                if in_spec:
                    # if None in line[first_index: last_index+1]: break
                    if len(line) != spec_line_cols_count or None in [line[i] for i in pattern]: continue  # игнорируем строки, набор столбцов которых не соответствует ранее зафиксированному набору для cпецификации
                    restored_cols = correct_row(line, pattern)
                    spec_line = [line[i] for i in pattern]
                    if all(cell == '' for cell in spec_line): continue  # Все столбцы содержат ''
                    if spec_line == ['1', '2', '3', '4', '5', '6', '7', '8', '9']: continue  # ['1', '2', '3', '4', '5', '6', '7', '8', '9'] игнорируем

                    
                    normalize_row(spec_line, prev_spec_line)
                    spec.append(spec_line)
                    spec_row_count += 1
                    if restored_cols:
                        restored_rows[spec_row_count] = restored_cols
                    prev_spec_line = spec_line
                    # spec.append(normalize_row(line[first_index: last_index+1]))
    if spec == []:
        raise ValueError("В файле спецификация не найдена")
    return spec, restored_rows


def detect_pattern(line):
    """Определяет шаблон таблицы спецификации по шапке спецификации (в которой все поля не пустые).
    input: line - список столбцов pdf-таблицы спецификации
    output: pattern - список индексов столбцов, содерщащих данные спецификации.
    """
    pattern = []
    for i, col in enumerate(line):
        if col == '' or col is None:
            continue
        pattern.append(i)
    return pattern
    # return [i for i, col in enumerate(line) if col not in ('', None)] # быстрее на 10–30%


def row_list_to_xlsx_bytes(spec, restored_rows, pdf_stem, template_path):
    """Формирует xlsx-файл спецификации на базе шаблона и возвращает его в виде байтов."""
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон спецификации не найден: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    if ws is None:
        raise ValueError("Не удалось получить активный лист из шаблона")

    # Записываем имя PDF в ячейку A1
    ws["A1"] = pdf_stem
    
    # 1. Сохраняем референсные стили из строки 3 (A3:I3)
    # Именно здесь в шаблоне настроены перенос текста, выравнивание и границы
    ref_styles = {}
    for col_idx in range(1, 10): # Столбцы A-I (индексы 1-9)
        ref_cell = ws.cell(row=3, column=col_idx)
        ref_styles[col_idx] = {
            'font': copy(ref_cell.font),
            'border': copy(ref_cell.border),
            'fill': copy(ref_cell.fill),
            'alignment': copy(ref_cell.alignment),
            'number_format': ref_cell.number_format
        }

    # 2. Пропускаем первую строку spec (так как это заголовок) 
    # и начинаем вывод данных с 3-й строки Excel
    data_rows = spec[1:] if len(spec) > 1 else []
    
    # 3. Заполняем данные и применяем сохраненные стили
    for row_offset, row_values in enumerate(data_rows):
        row_idx = 3 + row_offset
        for col_idx, value in enumerate(row_values, start=1):
            if col_idx > 9:
                break
                
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Применяем скопированные стили из эталонной строки
            styles = ref_styles.get(col_idx)
            if styles:
                cell.font = styles['font']
                cell.border = styles['border']
                cell.fill = styles['fill']
                cell.alignment = styles['alignment']
                cell.number_format = styles['number_format']

    # 4. Жестко фиксируем автофильтр на строке заголовков (строка 2)
    # Указываем диапазон до последней заполненной строки, чтобы фильтр не "уехал"
    last_row = 2 + len(data_rows)
    if last_row > 2:
        ws.auto_filter.ref = f"A2:I{last_row}"
    else:
        ws.auto_filter.ref = "A2:I2"

    # 5. Закрашиваем красным пустые ячейки в столбце "Наименование"
    for row_idx in range(3, last_row + 1):
        cell = ws.cell(row=row_idx, column=2)
        if not cell.value:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 6. Закрашиваем жёлтым ячейки, которые были восстановлены из "битой" табличной строки
    for row_idx, row_values in restored_rows.items():
        for col_idx in row_values:
            cell = ws.cell(row=row_idx+1, column=col_idx+1)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
