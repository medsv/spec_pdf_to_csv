import pymupdf4llm
import re
from copy import copy
from io import BytesIO
from openpyxl import load_workbook
from typing import List, Optional
import json

def pdf_spec_to_row_list(pdf_path):
    cols_count = 9
    spec_data = []
    prev_spec_line = None
    first_table = True
    
    try:
        json_str = pymupdf4llm.to_json(pdf_path)
        json_data = json.loads(json_str)
        
        for page in json_data.get('pages', []):
            for block in page.get('boxes', []):
                if block.get('boxclass') != 'table':
                    continue
                    
                json_table_data = block.get('table', {})
                if json_table_data.get('col_count', 0) != cols_count: 
                    continue
                if json_table_data.get('row_count', 0) < 2: 
                    continue
                    
                table_data = json_table_data.get('extract', [])
                if not table_data:
                    continue
                
                # Проверяем шапку мягко (игнорируя регистр и пробелы)
                header = table_data[0]
                if len(header) < cols_count:
                    continue
                if "примечание" not in str(header[8]).strip().lower():
                    continue
                    
                if first_table: 
                    spec_data.append(header)
                    first_table = False
                  
                for row in table_data[1:]:
                    if len(row) != cols_count:
                        continue
                    if all(cell == '' for cell in row):
                        continue
                        
                    # Игнорируем служебные строки нумерации колонок
                    if row == [str(i) for i in range(1, cols_count+1)]: 
                        continue
                        
                    # Обработка многострочных ячеек (продолжение строки)
                    is_section_header = bool(re.match(r'^\d+\.\d*', str(row[1]).strip()))
                    if row[0] == '' and row[1] != '' and not is_section_header and prev_spec_line is not None:
                        # Добавляем текст в "Примечание"
                        note_text = str(row[1]).strip()
                        if prev_spec_line[8]:
                            prev_spec_line[8] += f"\n{note_text}"
                        else:
                            prev_spec_line[8] = note_text
                        continue

                    spec_line = row.copy()
                    normalize_row(spec_line, prev_spec_line)
                    
                    spec_data.append(spec_line)
                    prev_spec_line = spec_line
                    
    except Exception as e:
        print(f"⚠️ Ошибка при обработке pdf-файла: {e}")    
        
    if not spec_data:
        raise ValueError("Спецификация не сформирована. Проверьте формат PDF.")
    return spec_data

def normalize_row(spec_line, prev_spec_line):
    """Очищает строку: удаляет лишние пробелы, нормализует единицы и др."""
    for i, col in enumerate(spec_line):
        if not isinstance(col, str):
            continue
        col = re.sub(r' +', ' ', col)
        col = col.strip()
        spec_line[i] = col
        
    # ставим пробел при ГОСТ12345-67
    if isinstance(spec_line[2], str):
        spec_line[2] = re.sub(r'ГОСТ(\d)', r'ГОСТ \1', spec_line[2])
        
    # Приводим единицу измерения к стандартному обозначению
    if isinstance(spec_line[5], str):
        spec_line[5] = normalize_unit(spec_line[5])
        
    # Замена "то же" наименованием из предыдущей строки
    if (
        prev_spec_line is not None
        and isinstance(spec_line[1], str)
        and spec_line[1].lower() == "то же"
        and prev_spec_line[1]
    ):
        spec_line[1] = prev_spec_line[1]

def row_list_to_xlsx_bytes(spec, pdf_stem, template_path):
    """Формирует xlsx-файл спецификации на базе шаблона и возвращает его в виде байтов."""
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон спецификации не найден: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    if ws is None:
        raise ValueError("Не удалось получить активный лист из шаблона")

    ws["A1"] = pdf_stem
    
    ref_styles = {}
    for col_idx in range(1, 10):
        ref_cell = ws.cell(row=3, column=col_idx)
        ref_styles[col_idx] = {
            'font': copy(ref_cell.font),
            'border': copy(ref_cell.border),
            'fill': copy(ref_cell.fill),
            'alignment': copy(ref_cell.alignment),
            'number_format': ref_cell.number_format
        }

    data_rows = spec[1:] if len(spec) > 1 else []
    
    for row_offset, row_values in enumerate(data_rows):
        row_idx = 3 + row_offset
        for col_idx, value in enumerate(row_values, start=1):
            if col_idx > 9:
                break
                
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Конвертация числовых полей для Excel (Кол-во - 7, Масса - 8)
            if col_idx in (7, 8) and value not in (None, ''):
                try:
                    # Меняем запятую на точку и преобразуем во float
                    cell.value = float(str(value).replace(',', '.'))
                except ValueError:
                    cell.value = value
            else:
                cell.value = value
            
            styles = ref_styles.get(col_idx)
            if styles:
                cell.font = styles['font']
                cell.border = styles['border']
                cell.fill = styles['fill']
                cell.alignment = styles['alignment']
                cell.number_format = styles['number_format']

    last_row = 2 + len(data_rows)
    ws.auto_filter.ref = f"A2:I{last_row}" if last_row > 2 else "A2:I2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def normalize_unit(unit: str) -> str:
    if not isinstance(unit, str): return unit
    unit = unit.lower().strip()
    unit = unit.replace('²', '2').replace('³', '3')
    cleaned = re.sub(r'[.\s]+', '', unit)
    
    MAPPING = {
        'м': 'м', 'metr': 'м', 'm': 'м',
        'м2': 'м2', 'квм': 'м2', 'sqm': 'м2',
        'м3': 'м3', 'кубм': 'м3', 'cum': 'м³',
        'л': 'л', 'l': 'л',
        'кг': 'кг', 'kg': 'кг',
        'т': 'т', 'тн': 'т', 't': 'т',
        'шт': 'шт.', 'штука': 'шт.',
        'уп': 'уп.', 
        'погм': 'пог. м', 'пм': 'пог. м', 'мп': 'пог. м',
        'компл': 'комплект',
    }
    return MAPPING.get(cleaned, unit)